import base64
import io
import unicodedata
from uuid import uuid4
from dataclasses import dataclass
from datetime import date, datetime
from typing import List, Dict, Any, Tuple, Optional

from app.menu.domain.monthly_menu import MonthlyMenu
from app.menu.domain.weekly_menu import WeeklyMenu
from app.menu.domain.daily_menu import DailyMenu
from app.menu.domain.meal import Meal
from app.menu.domain.meal_component import MealComponent
from app.menu.domain.menu_enums import MenuStatus, MealType
from app.menu.domain.component_type import ComponentType

from app.menu.application.ports.monthly_menu_repository import MonthlyMenuRepository
from app.menu.application.ports.weekly_menu_repository import WeeklyMenuRepository
from app.menu.application.ports.daily_menu_repository import DailyMenuRepository
from app.menu.application.ports.meal_repository import MealRepository
from app.menu.application.ports.meal_component_repository import MealComponentRepository
from app.menu.application.ports.component_type_repository import ComponentTypeRepository

try:
    # openpyxl nos permite leer la estructura de la hoja tal cual la ve el nutricionista
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


DAY_NAMES = {
    "LUNES",
    "MARTES",
    "MIERCOLES",
    "MIÉRCOLES",
    "JUEVES",
    "VIERNES",
    "SABADO",
    "SÁBADO",
    "DOMINGO",
}


@dataclass(frozen=True)
class UploadMonthlyMenuCommand:
    year: int
    month: int
    filename: str
    file_base64: str  # base64 de xlsx


class UploadMonthlyMenuUseCase:
    """
    Versión NORMALIZADA para el nuevo formato de Excel (4 hojas, desayuno/almuerzo/cena
    con múltiples componentes y kcal por día).

    - Lee cada hoja como una semana.
    - Detecta los días y fechas de la fila de cabecera (LUNES, MARTES, ...).
    - Para cada bloque (DESAYUNO, ALMUERZO, CENA) crea Meals.
    - Para cada fila de componente (BEBIDA CALIENTE, PLATO FONDO 1, etc.)
      crea MealComponents y usa component_types reales.
    """

    def __init__(
        self,
        monthly_repo: MonthlyMenuRepository,
        weekly_repo: WeeklyMenuRepository,
        daily_repo: DailyMenuRepository,
        meal_repo: MealRepository,
        meal_component_repo: MealComponentRepository,
        component_type_repo: ComponentTypeRepository,
    ) -> None:
        self.monthly_repo = monthly_repo
        self.weekly_repo = weekly_repo
        self.daily_repo = daily_repo
        self.meal_repo = meal_repo
        self.meal_component_repo = meal_component_repo
        self.component_type_repo = component_type_repo

        # cache in-memory para no pegarle a la BD por cada fila
        self._component_type_cache: Dict[str, ComponentType] = {}

    # =========================
    # Helpers de normalización
    # =========================
    @staticmethod
    def _decode_file(filename: str, file_base64: str) -> Tuple[bytes, str]:
        payload = base64.b64decode(file_base64)
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        return payload, ext

    @staticmethod
    def _normalize_str(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip().upper()
        # quitar acentos
        text = "".join(
            c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
        )
        return text

    @staticmethod
    def _clean_cell_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text in ("", "-----", "XXX", "####", "##"):
            return ""
        return text

    @staticmethod
    def _parse_kcal(value: Any) -> Optional[float]:
        if value is None:
            return None
        text = str(value).strip()
        if text in ("", "-----", "XXX", "####", "##", "TOTAL KCAL.", "TOTAL KCAL"):
            return None
        try:
            return float(text.replace(",", "."))
        except ValueError:
            return None

    async def _get_or_create_component_type(self, raw_label: str) -> ComponentType:
        """
        Devuelve o crea un tipo de componente.
        Aseguramos que NUNCA retorne un id inexistente.
        Limpieza automática de valores inválidos del Excel ("----", "##", etc.)
        """
        label = (raw_label or "").strip()

        # Normalización fuerte
        invalid = {"", "-", "----", "-----", "###", "##", "XXX"}
        if label.upper() in invalid:
            label = "GENÉRICO"

        # Cache (evita db roundtrips)
        cached = self._component_type_cache.get(label)
        if cached:
            return cached

        # Buscar en BD
        existing = await self.component_type_repo.get_by_name(label)
        if existing:
            self._component_type_cache[label] = existing
            return existing

        # Crear nuevo
        new_ct = ComponentType(id=None, name=label, display_order=0)
        created = await self.component_type_repo.create(new_ct)

        self._component_type_cache[label] = created
        return created

    # =========================
    # Parsing del Excel
    # =========================
    def _find_day_header_row(self, ws) -> Optional[int]:
        """
        Busca la fila donde aparecen LUNES, MARTES, etc.
        """
        for row in range(1, ws.max_row + 1):
            day_cells = 0
            for col in range(1, ws.max_column + 1):
                raw = ws.cell(row=row, column=col).value
                norm = self._normalize_str(raw)
                if norm in DAY_NAMES:
                    day_cells += 1
            if day_cells >= 2:
                return row
        return None

    def _extract_days_from_sheet(self, ws):
        """
        Lee correctamente días y fechas incluso con celdas combinadas.
        """
        result = []

        day_row = self._find_day_header_row(ws)
        if not day_row:
            return result

        date_row = day_row + 1

        col = 1
        while col <= ws.max_column:
            raw_day = ws.cell(day_row, col).value
            norm = self._normalize_str(raw_day)

            if norm in DAY_NAMES:
                # Corrección: búsqueda real del merge
                raw_date = ws.cell(date_row, col).value
                if raw_date is None:
                    # Buscar hacia la IZQUIERDA dentro del merge
                    cc = col
                    while cc > 1 and raw_date is None:
                        cc -= 1
                        raw_date = ws.cell(date_row, cc).value

                if not raw_date:
                    col += 2
                    continue

                # parseo seguro
                if isinstance(raw_date, datetime):
                    d = raw_date.date()
                elif isinstance(raw_date, date):
                    d = raw_date
                else:
                    d = datetime.strptime(str(raw_date), "%d/%m/%Y").date()

                result.append((d, str(raw_day).strip(), col, col + 1))
                col += 2
            else:
                col += 1

        return result


    def _find_blocks(self, ws) -> Dict[str, Tuple[int, int]]:
        """
        Detecta rangos de filas para DESAYUNO, ALMUERZO y CENA en función
        de la primera columna (donde salen los títulos de bloque).
        Devuelve dict:
            {
                "BREAKFAST": (row_start, row_end),
                "LUNCH": (row_start, row_end),
                "DINNER": (row_start, row_end),
            }
        """
        labels: Dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            raw = ws.cell(row=row, column=1).value
            norm = self._normalize_str(raw)
            if "DESAYUNO" in norm and "BREAKFAST" not in labels:
                labels["BREAKFAST"] = row
            elif "ALMUERZO" in norm and "LUNCH" not in labels:
                labels["LUNCH"] = row
            elif "CENA" in norm and "DINNER" not in labels:
                labels["DINNER"] = row

        blocks: Dict[str, Tuple[int, int]] = {}
        if "BREAKFAST" in labels:
            start = labels["BREAKFAST"] + 1
            end = labels.get("LUNCH") or labels.get("DINNER") or ws.max_row
            blocks["BREAKFAST"] = (start, end - 1)
        if "LUNCH" in labels:
            start = labels["LUNCH"] + 1
            end = labels.get("DINNER") or ws.max_row
            blocks["LUNCH"] = (start, end - 1)
        if "DINNER" in labels:
            start = labels["DINNER"] + 1
            end = ws.max_row
            blocks["DINNER"] = (start, end)

        return blocks

    def _extract_meal_components_for_day(
        self,
        ws,
        start_row: int,
        end_row: int,
        name_col: int,
        kcal_col: int,
    ) -> Tuple[List[Dict[str, Any]], Optional[float]]:
        """
        Lee un bloque (desayuno/almuerzo/cena) para un día concreto.
        Devuelve:
          - lista de dicts {component_label, dish_name, calories}
          - total_kcal (si existe fila TOTAL Kcal.)
        """
        components: List[Dict[str, Any]] = []
        total_kcal: Optional[float] = None

        for row in range(start_row, end_row + 1):
            label_raw = ws.cell(row=row, column=1).value
            label = self._clean_cell_text(label_raw)
            norm_label = self._normalize_str(label_raw)

            # detectar la fila TOTAL Kcal.
            if "TOTAL" in norm_label and "KCAL" in norm_label:
                total_kcal = self._parse_kcal(ws.cell(row=row, column=kcal_col).value)
                continue

            if not label:
                # fila vacía / separador
                continue

            dish_name = self._clean_cell_text(ws.cell(row=row, column=name_col).value)
            calories = self._parse_kcal(ws.cell(row=row, column=kcal_col).value)

            if not dish_name:
                # no hay plato para este componente en este día
                continue

            components.append(
                {
                    "component_label": label,
                    "dish_name": dish_name,
                    "calories": calories,
                }
            )

        return components, total_kcal

    # =========================
    # Ejecución principal
    # =========================
    async def execute(self, cmd: UploadMonthlyMenuCommand) -> Dict[str, Any]:
        # 1) upsert MonthlyMenu
        monthly = await self.monthly_repo.find_by_year_month(cmd.year, cmd.month)
        if not monthly:
            monthly = MonthlyMenu(
                id=str(uuid4()),
                year=cmd.year,
                month=cmd.month,
                status=MenuStatus.DRAFT,
                source_filename=cmd.filename,
            )
        else:
            monthly.source_filename = cmd.filename
        monthly = await self.monthly_repo.upsert(monthly)

        # 2) decodificar archivo y cargar workbook
        payload, ext = self._decode_file(cmd.filename, cmd.file_base64)
        if ext not in {"xlsx", "xls"}:
            raise ValueError("Solo se soportan archivos Excel (.xlsx, .xls) para el nuevo formato de menú.")

        if load_workbook is None:
            raise RuntimeError(
                "Se envió un Excel de menú pero no está instalado 'openpyxl' en el servidor."
            )

        wb = load_workbook(io.BytesIO(payload), data_only=True)

        # 3) Crear las semanas (una por hoja)
        sheetnames = wb.sheetnames
        weeks: List[WeeklyMenu] = []
        for idx, sheet_name in enumerate(sheetnames, start=1):
            weeks.append(
                WeeklyMenu(
                    id=None,
                    monthly_menu_id=str(monthly.id),
                    week_number=idx,
                    title=sheet_name,
                )
            )

        await self.weekly_repo.bulk_replace(str(monthly.id), weeks)
        persisted_weeks = await self.weekly_repo.list_by_month(str(monthly.id))
        weeks_by_number = {w.week_number: w for w in persisted_weeks}

        # 4) Por cada hoja (semana) procesar días, comidas y componentes
        for idx, sheet_name in enumerate(sheetnames, start=1):
            ws = wb[sheet_name]
            weekly = weeks_by_number.get(idx)
            if not weekly:
                continue

            # detectar días/fechas de la cabecera
            days_info = self._extract_days_from_sheet(ws)
            # filtrar solo fechas del mes/año indicado
            days_info = [
                item
                for item in days_info
                if item[0].year == cmd.year and item[0].month == cmd.month
            ]
            if not days_info:
                # semana sin días de este mes
                await self.daily_repo.bulk_replace_for_week(str(weekly.id), [])
                continue

            # crear DailyMenus para la semana
            daily_menus: List[DailyMenu] = []
            for d, day_name, _name_col, _kcal_col in days_info:
                daily_menus.append(
                    DailyMenu(
                        id=None,
                        weekly_menu_id=str(weekly.id),
                        date=d,
                        day_of_week=day_name.upper(),
                        is_holiday=False,
                    )
                )

            await self.daily_repo.bulk_replace_for_week(str(weekly.id), daily_menus)
            persisted_days = await self.daily_repo.list_by_week(str(weekly.id))
            day_by_date = {dm.date: dm for dm in persisted_days}

            # localizar bloques de desayuno/almuerzo/cena en la hoja
            blocks = self._find_blocks(ws)

            # para cada día, crear meals y meal_components
            for d, _day_name, name_col, kcal_col in days_info:
                day_menu = day_by_date.get(d)
                if not day_menu:
                    continue

                meals_to_create: List[Meal] = []
                components_by_meal_type: Dict[MealType, List[Dict[str, Any]]] = {}

                # DESAYUNO
                if "BREAKFAST" in blocks:
                    comps, total_kcal = self._extract_meal_components_for_day(
                        ws,
                        start_row=blocks["BREAKFAST"][0],
                        end_row=blocks["BREAKFAST"][1],
                        name_col=name_col,
                        kcal_col=kcal_col,
                    )
                    if comps or total_kcal is not None:
                        meals_to_create.append(
                            Meal(
                                id=None,
                                daily_menu_id=str(day_menu.id),
                                meal_type=MealType.BREAKFAST,
                                total_kcal=total_kcal,
                            )
                        )
                        components_by_meal_type[MealType.BREAKFAST] = comps

                # ALMUERZO
                if "LUNCH" in blocks:
                    comps, total_kcal = self._extract_meal_components_for_day(
                        ws,
                        start_row=blocks["LUNCH"][0],
                        end_row=blocks["LUNCH"][1],
                        name_col=name_col,
                        kcal_col=kcal_col,
                    )
                    if comps or total_kcal is not None:
                        meals_to_create.append(
                            Meal(
                                id=None,
                                daily_menu_id=str(day_menu.id),
                                meal_type=MealType.LUNCH,
                                total_kcal=total_kcal,
                            )
                        )
                        components_by_meal_type[MealType.LUNCH] = comps

                # CENA
                if "DINNER" in blocks:
                    comps, total_kcal = self._extract_meal_components_for_day(
                        ws,
                        start_row=blocks["DINNER"][0],
                        end_row=blocks["DINNER"][1],
                        name_col=name_col,
                        kcal_col=kcal_col,
                    )
                    if comps or total_kcal is not None:
                        meals_to_create.append(
                            Meal(
                                id=None,
                                daily_menu_id=str(day_menu.id),
                                meal_type=MealType.DINNER,
                                total_kcal=total_kcal,
                            )
                        )
                        components_by_meal_type[MealType.DINNER] = comps

                # reemplazar comidas del día
                await self.meal_repo.bulk_replace_for_daily_menu(str(day_menu.id), meals_to_create)
                persisted_meals = await self.meal_repo.list_by_daily_menu(str(day_menu.id))

                # para cada meal persistido, crear sus componentes
                for meal in persisted_meals:
                    comps_data = components_by_meal_type.get(meal.meal_type, [])
                    components: List[MealComponent] = []
                    order_position = 0

                    for comp in comps_data:
                        label = comp["component_label"]
                        dish_name = comp["dish_name"]
                        calories = comp["calories"]

                        if not dish_name:
                            continue

                        component_type = await self._get_or_create_component_type(label)
                        order_position += 1

                        components.append(
                            MealComponent(
                                id=None,
                                meal_id=str(meal.id),
                                component_type_id=str(component_type.id),
                                dish_name=dish_name,
                                calories=calories,
                                order_position=order_position,
                            )
                        )

                    await self.meal_component_repo.bulk_replace_for_meal(str(meal.id), components)

        # 5) activar el menú mensual
        monthly.activate()
        await self.monthly_repo.upsert(monthly)

        return {"status": "ok", "message": "Menú mensual cargado correctamente."}
