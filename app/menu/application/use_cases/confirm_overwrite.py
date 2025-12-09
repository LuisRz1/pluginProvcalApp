from dataclasses import dataclass
from typing import Dict, Any, List, Tuple, Optional
import base64
import io
from datetime import date, datetime

from app.menu.application.ports.monthly_menu_repository import MonthlyMenuRepository
from app.menu.application.ports.weekly_menu_repository import WeeklyMenuRepository
from app.menu.application.ports.daily_menu_repository import DailyMenuRepository
from app.menu.application.ports.meal_repository import MealRepository
from app.menu.application.ports.meal_component_repository import MealComponentRepository

# Este UC solo PREVISA el contenido detectado; no escribe en BD.
# Se usa antes de hacer el upload definitivo para validar el archivo.

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


@dataclass(frozen=True)
class ConfirmOverwriteCommand:
    year: int
    month: int
    filename: str
    file_base64: str  # base64 del Excel


class ConfirmOverwriteUseCase:
    def __init__(
        self,
        monthly_repo: MonthlyMenuRepository,
        weekly_repo: WeeklyMenuRepository,
        daily_repo: DailyMenuRepository,
        meal_repo: MealRepository,
        meal_component_repo: MealComponentRepository,
    ) -> None:
        # Se inyectan por compatibilidad, aunque este UC no escribe en BD.
        self.monthly_repo = monthly_repo
        self.weekly_repo = weekly_repo
        self.daily_repo = daily_repo
        self.meal_repo = meal_repo
        self.meal_component_repo = meal_component_repo

    # ============
    # Helpers
    # ============
    @staticmethod
    def _decode_file(filename: str, file_base64: str) -> Tuple[bytes, str]:
        payload = base64.b64decode(file_base64)
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        return payload, ext

    @staticmethod
    def _normalize_str(value) -> str:
        if value is None:
            return ""
        return str(value).strip().upper()

    def _find_day_header_row(self, ws) -> Optional[int]:
        """
        Busca la fila donde aparecen LUNES, MARTES, etc.
        """
        from app.menu.application.use_cases.upload_monthly_menu import DAY_NAMES  # evitar duplicar constantes

        for row in range(1, ws.max_row + 1):
            day_cells = 0
            for col in range(1, ws.max_column + 1):
                raw = ws.cell(row=row, column=col).value
                norm = self._normalize_str(raw)
                if norm in DAY_NAMES:
                    day_cells += 1
            if day_cells >= 2:
                return row
        return None

    def _extract_days_from_sheet(self, ws) -> List[Tuple[date, str]]:
        """
        Devuelve lista de (fecha, nombre_dia) detectados en la hoja.
        """
        result: List[Tuple[date, str]] = []
        day_row = self._find_day_header_row(ws)
        if not day_row:
            return result
        date_row = day_row + 1

        from app.menu.application.use_cases.upload_monthly_menu import DAY_NAMES

        col = 1
        while col <= ws.max_column:
            raw_day = ws.cell(row=day_row, column=col).value
            norm_day = self._normalize_str(raw_day)

            if norm_day in DAY_NAMES:
                raw_date = ws.cell(row=date_row, column=col).value
                if not raw_date:
                    col += 2
                    continue

                if isinstance(raw_date, datetime):
                    d = raw_date.date()
                elif isinstance(raw_date, date):
                    d = raw_date
                else:
                    text = str(raw_date).strip()
                    parsed: Optional[date] = None
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
                        try:
                            parsed = datetime.strptime(text, fmt).date()
                            break
                        except ValueError:
                            continue
                    if not parsed:
                        col += 2
                        continue
                    d = parsed

                result.append((d, str(raw_day).strip()))
                col += 2
            else:
                col += 1

        return result

    # ============
    # Execute
    # ============
    async def execute(self, cmd: ConfirmOverwriteCommand) -> Dict[str, Any]:
        payload, ext = self._decode_file(cmd.filename, cmd.file_base64)

        if ext not in {"xlsx", "xls"}:
            return {
                "status": "error",
                "message": "El archivo debe ser un Excel (.xlsx o .xls) con el formato de menú mensual.",
                "preview": {},
            }

        if load_workbook is None:
            return {
                "status": "error",
                "message": "No está instalado 'openpyxl' en el servidor para leer archivos Excel.",
                "preview": {},
            }

        wb = load_workbook(io.BytesIO(payload), data_only=True)

        sheets_preview: List[Dict[str, Any]] = []
        total_days = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            days = self._extract_days_from_sheet(ws)
            # Filtrar solo días del mes/año que se quiere cargar
            filtered = [
                (d, name)
                for d, name in days
                if d.year == cmd.year and d.month == cmd.month
            ]

            total_days += len(filtered)
            sheets_preview.append(
                {
                    "sheet": sheet_name,
                    "day_count": len(filtered),
                    "days": [
                        {
                            "date": d.isoformat(),
                            "day_name": name,
                        }
                        for d, name in filtered
                    ],
                }
            )

        if total_days == 0:
            status = "warning"
            message = (
                f"No se encontraron días del mes {cmd.month:02d}/{cmd.year} en el archivo. "
                "Verifica que el mes/año coincidan y que las cabeceras de días estén bien (LUNES, MARTES, etc.)."
            )
        else:
            status = "ok"
            message = (
                f"Archivo reconocido. Se encontraron {total_days} días del mes {cmd.month:02d}/{cmd.year} "
                "distribuidos en las distintas hojas. Listo para subir."
            )

        preview: Dict[str, Any] = {
            "year": cmd.year,
            "month": cmd.month,
            "total_days": total_days,
            "sheets": sheets_preview,
        }

        return {
            "status": status,
            "message": message,
            "preview": preview,
        }
